#!/usr/bin/env python3
"""
SCRMS-PDCA: Control Maturity Assessment Generator
Generates a structured Excel workbook for CMM-based control maturity assessment.

Usage:
    python generate_assessment.py --output <path> [--org-name <name>] [--domains <d1,d2,...>]

Default domains (if not specified):
    Access Management, Vulnerability Management, Incident Response,
    Data Protection, Third-Party Risk, Security Awareness,
    Change Management, Business Continuity, Physical Security, Logging & Monitoring
"""

import argparse
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation


# ── Color palette ──────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":    "1F3864",   # dark navy
    "header_fg":    "FFFFFF",
    "section_bg":   "2E75B6",   # medium blue
    "section_fg":   "FFFFFF",
    "subheader_bg": "D6E4F7",   # light blue
    "cmm0":         "FF0000",   # red
    "cmm1":         "FF6600",   # orange
    "cmm2":         "FFC000",   # amber
    "cmm3":         "92D050",   # yellow-green
    "cmm4":         "00B050",   # green
    "cmm5":         "00B0F0",   # cyan
    "gap_high":     "FFE0E0",
    "gap_med":      "FFF3CD",
    "gap_low":      "E8F5E9",
    "alt_row":      "F2F7FB",
    "border":       "B8CCE4",
}

DEFAULT_DOMAINS = [
    "Access Management",
    "Vulnerability Management",
    "Incident Response",
    "Data Protection & Privacy",
    "Third-Party Risk Management",
    "Security Awareness & Training",
    "Change Management",
    "Business Continuity & DR",
    "Physical & Environmental Security",
    "Logging, Monitoring & SIEM",
]

SAMPLE_CONTROLS = {
    "Access Management": [
        ("AC-1",  "Access Control Policy & Procedures",                "MCR"),
        ("AC-2",  "Account Management (provisioning/de-provisioning)", "MCR"),
        ("AC-3",  "Access Enforcement (least privilege)",              "MCR"),
        ("AC-6",  "Privileged Access Management (PAM)",                "MCR"),
        ("AC-11", "Multi-Factor Authentication (MFA)",                  "MCR"),
        ("AC-17", "Remote Access Controls",                            "MCR"),
        ("AC-20", "Periodic Access Reviews / Recertification",         "DSR"),
    ],
    "Vulnerability Management": [
        ("VM-1",  "Vulnerability Scanning (internal)",                 "MCR"),
        ("VM-2",  "Vulnerability Scanning (external / perimeter)",     "MCR"),
        ("VM-3",  "Patch Management Process",                          "MCR"),
        ("VM-4",  "Vulnerability Prioritization & SLA",                "MCR"),
        ("VM-5",  "Penetration Testing Program",                       "DSR"),
        ("VM-6",  "Vulnerability Metrics & Trending",                  "DSR"),
    ],
    "Incident Response": [
        ("IR-1",  "Incident Response Plan (IRP)",                     "MCR"),
        ("IR-2",  "Incident Detection & Analysis",                    "MCR"),
        ("IR-3",  "Incident Containment & Eradication",               "MCR"),
        ("IR-4",  "Post-Incident Review / Lessons Learned",           "MCR"),
        ("IR-5",  "Tabletop Exercises / IR Testing",                  "DSR"),
        ("IR-6",  "Threat Intelligence Integration",                   "DSR"),
    ],
    "Data Protection & Privacy": [
        ("DP-1",  "Data Classification Policy",                       "MCR"),
        ("DP-2",  "Data Inventory / Records of Processing",           "MCR"),
        ("DP-3",  "Encryption at Rest",                               "MCR"),
        ("DP-4",  "Encryption in Transit",                            "MCR"),
        ("DP-5",  "Data Retention & Disposal",                        "MCR"),
        ("DP-6",  "Data Loss Prevention (DLP)",                       "DSR"),
        ("DP-7",  "Privacy Impact Assessments (PIA/DPIA)",            "DSR"),
    ],
    "Third-Party Risk Management": [
        ("TPR-1", "Vendor Risk Assessment Process",                   "MCR"),
        ("TPR-2", "Third-Party Security Requirements (contracts)",    "MCR"),
        ("TPR-3", "Vendor Tiering / Risk Classification",             "MCR"),
        ("TPR-4", "Ongoing Vendor Monitoring",                        "DSR"),
        ("TPR-5", "Fourth-Party / Supply Chain Risk",                 "DSR"),
    ],
    "Security Awareness & Training": [
        ("SAT-1", "Annual Security Awareness Training",               "MCR"),
        ("SAT-2", "Role-Based Security Training",                     "MCR"),
        ("SAT-3", "Phishing Simulation Program",                      "DSR"),
        ("SAT-4", "Privacy Training for Data Handlers",               "MCR"),
    ],
    "Change Management": [
        ("CM-1",  "Change Management Policy & Procedures",            "MCR"),
        ("CM-2",  "Change Advisory Board (CAB) Process",              "MCR"),
        ("CM-3",  "Emergency Change Procedures",                      "MCR"),
        ("CM-4",  "Configuration Baseline Management",                "MCR"),
        ("CM-5",  "Configuration Drift Detection",                    "DSR"),
    ],
    "Business Continuity & DR": [
        ("BC-1",  "Business Continuity Plan (BCP)",                   "MCR"),
        ("BC-2",  "Disaster Recovery Plan (DRP)",                     "MCR"),
        ("BC-3",  "Business Impact Analysis (BIA)",                   "MCR"),
        ("BC-4",  "BCP/DRP Testing (tabletop & full exercise)",       "MCR"),
        ("BC-5",  "Recovery Time / Point Objectives (RTO/RPO)",       "MCR"),
        ("BC-6",  "Backup Validation & Testing",                      "DSR"),
    ],
    "Physical & Environmental Security": [
        ("PE-1",  "Physical Access Controls (data centers)",          "MCR"),
        ("PE-2",  "Visitor Management",                               "MCR"),
        ("PE-3",  "Environmental Monitoring (temp, humidity, power)", "DSR"),
        ("PE-4",  "Secure Disposal of Physical Media",                "MCR"),
    ],
    "Logging, Monitoring & SIEM": [
        ("LM-1",  "Audit Logging Policy (what to log)",               "MCR"),
        ("LM-2",  "Log Aggregation & SIEM",                           "MCR"),
        ("LM-3",  "Log Retention (meets compliance requirements)",    "MCR"),
        ("LM-4",  "Security Event Alerting & Triage",                 "MCR"),
        ("LM-5",  "Log Integrity Protection",                         "MCR"),
        ("LM-6",  "User & Entity Behavior Analytics (UEBA)",          "DSR"),
    ],
}

CMM_LABELS = {
    0: "CMM 0 – Not Performed",
    1: "CMM 1 – Performed Informally",
    2: "CMM 2 – Planned & Tracked",
    3: "CMM 3 – Well-Defined",
    4: "CMM 4 – Quantitatively Controlled",
    5: "CMM 5 – Continuously Improving",
}

CMM_FILL_KEYS = ["cmm0", "cmm1", "cmm2", "cmm3", "cmm4", "cmm5"]


def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size)

def make_border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=False):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def style_header_cell(cell, text, bg=None, fg=None, size=11, bold=True, center=True):
    cell.value = text
    cell.fill = make_fill(bg or COLORS["header_bg"])
    cell.font = make_font(bold=bold, color=fg or COLORS["header_fg"], size=size)
    cell.alignment = center_align(wrap=True) if center else left_align(wrap=True)
    cell.border = make_border()


def build_assessment_sheet(wb, org_name, domains):
    ws = wb.active
    ws.title = "Maturity Assessment"

    # ── Title row ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"SCRMS-PDCA Control Maturity Assessment — {org_name}"
    title_cell.fill = make_fill(COLORS["header_bg"])
    title_cell.font = make_font(bold=True, color=COLORS["header_fg"], size=14)
    title_cell.alignment = center_align()
    ws.row_dimensions[1].height = 30

    # ── Subtitle row ──────────────────────────────────────────────────────────
    ws.merge_cells("A2:M2")
    sub = ws["A2"]
    sub.value = f"Assessment Date: {date.today().strftime('%B %d, %Y')}   |   Framework: SCF / SCRMS v2026.1   |   PDCA Phase: PLAN → Principle #3"
    sub.fill = make_fill(COLORS["section_bg"])
    sub.font = make_font(bold=False, color=COLORS["header_fg"], size=10)
    sub.alignment = center_align()
    ws.row_dimensions[2].height = 18

    # ── CMM Legend ─────────────────────────────────────────────────────────────
    ws.merge_cells("A3:M3")
    ws["A3"].value = "CMM Legend:"
    ws["A3"].font = make_font(bold=True, size=10)
    ws["A3"].alignment = left_align()
    ws.row_dimensions[3].height = 16

    legend_row = 4
    for i, (level, label) in enumerate(CMM_LABELS.items()):
        col = i * 2 + 1
        cell = ws.cell(row=legend_row, column=col)
        cell.value = label
        cell.fill = make_fill(COLORS[CMM_FILL_KEYS[level]])
        cell.font = make_font(bold=True, color="FFFFFF" if level in (0, 1) else "000000", size=9)
        cell.alignment = center_align()
        cell.border = make_border()
        ws.merge_cells(start_row=legend_row, start_column=col, end_row=legend_row, end_column=col + 1)
    ws.row_dimensions[legend_row].height = 20

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = [
        ("A", "Control ID",          10),
        ("B", "Control Name",        40),
        ("C", "MCR / DSR",            9),
        ("D", "Current CMM\n(0–5)",  12),
        ("E", "Target CMM\n(0–5)",   12),
        ("F", "Gap",                  7),
        ("G", "Priority\n(H/M/L)",   10),
        ("H", "Control Owner",       22),
        ("I", "Control Operator",    22),
        ("J", "Evidence of\nCurrent Level", 28),
        ("K", "Improvement Actions", 35),
        ("L", "Target Date",         13),
        ("M", "Notes",               28),
    ]

    header_row = 5
    ws.row_dimensions[header_row].height = 36
    for col_letter, label, width in headers:
        cell = ws[f"{col_letter}{header_row}"]
        style_header_cell(cell, label)
        ws.column_dimensions[col_letter].width = width

    # ── Data validation for CMM columns ───────────────────────────────────────
    cmm_dv = DataValidation(type="whole", operator="between", formula1="0", formula2="5",
                            showErrorMessage=True, errorTitle="Invalid CMM",
                            error="Enter a value from 0 (Not Performed) to 5 (Continuously Improving)")
    ws.add_data_validation(cmm_dv)

    priority_dv = DataValidation(type="list", formula1='"High,Medium,Low"',
                                 showErrorMessage=True)
    ws.add_data_validation(priority_dv)

    mcr_dv = DataValidation(type="list", formula1='"MCR,DSR"', showErrorMessage=True)
    ws.add_data_validation(mcr_dv)

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_row = header_row + 1
    alt = False

    for domain in domains:
        # Section header
        ws.merge_cells(f"A{current_row}:M{current_row}")
        sec = ws[f"A{current_row}"]
        sec.value = f"  ▶  {domain}"
        sec.fill = make_fill(COLORS["section_bg"])
        sec.font = make_font(bold=True, color=COLORS["header_fg"], size=11)
        sec.alignment = left_align()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        controls = SAMPLE_CONTROLS.get(domain, [
            ("TBD-1", "Define controls for this domain", "MCR"),
        ])

        for ctrl_id, ctrl_name, ctrl_type in controls:
            row_bg = COLORS["alt_row"] if alt else "FFFFFF"
            alt = not alt

            data = [
                ("A", ctrl_id,    False),
                ("B", ctrl_name,  False),
                ("C", ctrl_type,  True),
                ("D", "",         True),   # Current CMM
                ("E", "",         True),   # Target CMM
                ("F", "",         True),   # Gap (manual or formula)
                ("G", "",         True),   # Priority
                ("H", "",         False),  # Owner
                ("I", "",         False),  # Operator
                ("J", "",         False),  # Evidence
                ("K", "",         False),  # Actions
                ("L", "",         True),   # Date
                ("M", "",         False),  # Notes
            ]

            for col_letter, value, centered in data:
                cell = ws[f"{col_letter}{current_row}"]
                cell.value = value
                cell.fill = make_fill(row_bg)
                cell.font = make_font(size=10)
                cell.alignment = center_align() if centered else left_align(wrap=True)
                cell.border = make_border()

            # Gap formula: =IF(AND(D>0,E>0), E-D, "")
            gap_cell = ws[f"F{current_row}"]
            d_ref = f"D{current_row}"
            e_ref = f"E{current_row}"
            gap_cell.value = f'=IF(AND({d_ref}<>"",{e_ref}<>""),{e_ref}-{d_ref},"")'

            # Add CMM data validations
            cmm_dv.add(ws[f"D{current_row}"])
            cmm_dv.add(ws[f"E{current_row}"])
            priority_dv.add(ws[f"G{current_row}"])
            mcr_dv.add(ws[f"C{current_row}"])

            ws.row_dimensions[current_row].height = 20
            current_row += 1

        # Spacer
        ws.row_dimensions[current_row].height = 6
        current_row += 1

    # Freeze panes below headers
    ws.freeze_panes = f"A{header_row + 1}"


def build_summary_sheet(wb, org_name, domains):
    ws = wb.create_sheet("Executive Summary")

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"SCRMS Maturity Summary — {org_name}"
    ws["A1"].fill = make_fill(COLORS["header_bg"])
    ws["A1"].font = make_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = center_align()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        "This sheet provides a high-level view of maturity gaps by domain. "
        "Populate the 'Maturity Assessment' sheet first."
    )
    ws["A2"].fill = make_fill(COLORS["subheader_bg"])
    ws["A2"].font = make_font(size=10)
    ws["A2"].alignment = left_align(wrap=True)
    ws.row_dimensions[2].height = 28

    headers = [
        ("A", "Domain",              38),
        ("B", "Avg Current CMM",     18),
        ("C", "Avg Target CMM",      18),
        ("D", "Avg Gap",             12),
        ("E", "# Controls",          12),
        ("F", "Priority",            14),
    ]
    for col_letter, label, width in headers:
        style_header_cell(ws[f"{col_letter}4"], label)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[4].height = 24

    for i, domain in enumerate(domains):
        row = i + 5
        ws[f"A{row}"].value = domain
        ws[f"A{row}"].font = make_font(size=10)
        ws[f"A{row}"].alignment = left_align()
        ws[f"A{row}"].border = make_border()
        for col in "BCDEF":
            cell = ws[f"{col}{row}"]
            cell.value = "(populate from assessment)"
            cell.font = make_font(size=9, color="888888")
            cell.alignment = center_align()
            cell.border = make_border()
        ws.row_dimensions[row].height = 18


def build_risk_register_sheet(wb):
    ws = wb.create_sheet("Risk Register (Principle 8)")

    ws.merge_cells("A1:K1")
    ws["A1"].value = "SCRMS Risk Register — Principle #8: Manage Risk"
    ws["A1"].fill = make_fill(COLORS["header_bg"])
    ws["A1"].font = make_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = center_align()
    ws.row_dimensions[1].height = 28

    headers = [
        ("A", "Risk ID",             10),
        ("B", "Risk Description",    38),
        ("C", "Source\n(Finding/Metric/Threat)", 20),
        ("D", "Likelihood\n(1–5)",   12),
        ("E", "Impact\n(1–5)",       11),
        ("F", "Inherent\nRisk Score", 12),
        ("G", "Treatment\n(R/A/T/Ac)", 14),
        ("H", "Control Owner",       22),
        ("I", "Residual Risk\nNotes", 28),
        ("J", "Target Date",         13),
        ("K", "Status",              14),
    ]

    for col_letter, label, width in headers:
        style_header_cell(ws[f"{col_letter}2"], label)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[2].height = 36

    treatment_dv = DataValidation(
        type="list",
        formula1='"Reduce,Avoid,Transfer,Accept"',
        showErrorMessage=True,
        errorTitle="Invalid Treatment",
        error="Select: Reduce, Avoid, Transfer, or Accept"
    )
    ws.add_data_validation(treatment_dv)

    status_dv = DataValidation(
        type="list",
        formula1='"Open,In Progress,Resolved,Accepted"',
        showErrorMessage=True
    )
    ws.add_data_validation(status_dv)

    for row in range(3, 23):
        row_bg = COLORS["alt_row"] if row % 2 == 0 else "FFFFFF"
        for col_letter in "ABCDEFGHIJK":
            cell = ws[f"{col_letter}{row}"]
            cell.fill = make_fill(row_bg)
            cell.border = make_border()
            cell.font = make_font(size=10)
            cell.alignment = left_align(wrap=True)

        # Inherent risk formula
        ws[f"F{row}"].value = f"=IF(AND(D{row}<>\"\",E{row}<>\"\"),D{row}*E{row},\"\")"
        ws[f"F{row}"].alignment = center_align()

        treatment_dv.add(ws[f"G{row}"])
        status_dv.add(ws[f"K{row}"])
        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A3"

    # Treatment legend
    legend_row = 24
    ws.merge_cells(f"A{legend_row}:K{legend_row}")
    ws[f"A{legend_row}"].value = (
        "Treatment Options: R = Reduce (implement/improve controls)  |  "
        "A = Avoid (stop the activity)  |  T = Transfer (insurance/contract)  |  Ac = Accept (document & sign off)"
    )
    ws[f"A{legend_row}"].font = make_font(size=9, color="444444")
    ws[f"A{legend_row}"].alignment = left_align()
    ws[f"A{legend_row}"].fill = make_fill(COLORS["subheader_bg"])


def build_metrics_sheet(wb):
    ws = wb.create_sheet("KPI-KRI-KCI Metrics (P7)")

    ws.merge_cells("A1:I1")
    ws["A1"].value = "SCRMS Metrics Framework — Principle #7: Maintain Situational Awareness"
    ws["A1"].fill = make_fill(COLORS["header_bg"])
    ws["A1"].font = make_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = center_align()
    ws.row_dimensions[1].height = 28

    headers = [
        ("A", "Metric ID",          10),
        ("B", "Metric Name",        35),
        ("C", "Type\n(KPI/KRI/KCI)", 12),
        ("D", "Formula / Data Source", 30),
        ("E", "Frequency",          14),
        ("F", "Target / Threshold", 22),
        ("G", "Owner",              20),
        ("H", "Linked Control",     20),
        ("I", "Audience",           18),
    ]

    for col_letter, label, width in headers:
        style_header_cell(ws[f"{col_letter}2"], label)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[2].height = 36

    sample_metrics = [
        ("KPI-001", "Mean Time to Patch (Critical CVEs)",       "KPI", "Avg days from CVE publish to patch applied", "Weekly",   "≤ 7 days",  "Vulnerability Mgr",  "VM-3", "Ops, CISO"),
        ("KPI-002", "Incident MTTD",                            "KPI", "Avg time from event to detection (hours)",   "Monthly",  "≤ 4 hours", "SOC Lead",           "IR-2", "CISO, Board"),
        ("KPI-003", "Security Training Completion Rate",        "KPI", "# completed / # assigned × 100",            "Monthly",  "≥ 95%",     "HR / Sec Awareness", "SAT-1","HR, CISO"),
        ("KRI-001", "Critical CVEs > 30 Days Unpatched",        "KRI", "Count of CVEs > SLA",                        "Weekly",   "= 0",       "Vulnerability Mgr",  "VM-4", "CISO"),
        ("KRI-002", "Privileged Accounts Without MFA",          "KRI", "Count of PAM accounts lacking MFA",          "Daily",    "= 0",       "IAM Team",           "AC-11","CISO, IAM Lead"),
        ("KRI-003", "Failed Auth Attempts (7-day trend)",       "KRI", "Count trend (↑ = early warning)",            "Daily",    "< baseline","SOC Analyst",        "LM-4", "SOC, CISO"),
        ("KCI-001", "Endpoints Meeting Hardening Baseline (%)", "KCI", "# compliant endpoints / total × 100",       "Weekly",   "≥ 98%",     "Endpoint Team",      "CM-4", "CISO, Ops"),
        ("KCI-002", "Data Classified per Policy (%)",           "KCI", "Classified assets / total assets × 100",    "Monthly",  "≥ 90%",     "Data Owner",         "DP-1", "Privacy, CISO"),
        ("KCI-003", "Third-Party Assessments Current (%)",      "KCI", "Vendors with current assessment / total",   "Quarterly","≥ 95%",     "Vendor Risk Mgr",    "TPR-1","Procurement, CISO"),
    ]

    type_dv = DataValidation(type="list", formula1='"KPI,KRI,KCI"', showErrorMessage=True)
    ws.add_data_validation(type_dv)

    for i, row_data in enumerate(sample_metrics):
        row = i + 3
        row_bg = COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
        for j, (col_letter, _) in enumerate(headers):
            cell = ws[f"{col_letter}{row}"]
            cell.value = row_data[j] if j < len(row_data) else ""
            cell.fill = make_fill(row_bg)
            cell.border = make_border()
            cell.font = make_font(size=10)
            cell.alignment = center_align() if col_letter in "ACE" else left_align(wrap=True)
        type_dv.add(ws[f"C{row}"])
        ws.row_dimensions[row].height = 22

    # Blank rows for additions
    for row in range(len(sample_metrics) + 3, len(sample_metrics) + 13):
        row_bg = COLORS["alt_row"] if row % 2 == 0 else "FFFFFF"
        for col_letter, _, _ in headers:
            cell = ws[f"{col_letter}{row}"]
            cell.fill = make_fill(row_bg)
            cell.border = make_border()
            cell.font = make_font(size=10)
            type_dv.add(ws[f"C{row}"])
        ws.row_dimensions[row].height = 22

    ws.freeze_panes = "A3"


def main():
    parser = argparse.ArgumentParser(
        description="Generate a SCRMS-PDCA Control Maturity Assessment workbook"
    )
    parser.add_argument("--output", default="SCRMS_Maturity_Assessment.xlsx",
                        help="Output file path (default: SCRMS_Maturity_Assessment.xlsx)")
    parser.add_argument("--org-name", default="[Organization Name]",
                        help="Organization name for the workbook header")
    parser.add_argument("--domains", default=None,
                        help="Comma-separated list of control domains (default: all 10)")
    args = parser.parse_args()

    domains = args.domains.split(",") if args.domains else DEFAULT_DOMAINS
    domains = [d.strip() for d in domains]

    print(f"Generating SCRMS Maturity Assessment for: {args.org_name}")
    print(f"Domains: {', '.join(domains)}")

    wb = openpyxl.Workbook()

    build_assessment_sheet(wb, args.org_name, domains)
    build_summary_sheet(wb, args.org_name, domains)
    build_risk_register_sheet(wb)
    build_metrics_sheet(wb)

    wb.save(args.output)
    print(f"\n✅ Workbook saved to: {args.output}")
    print("\nSheets created:")
    print("  1. Maturity Assessment  — CMM scoring per control (Principle #3)")
    print("  2. Executive Summary    — Domain-level maturity overview")
    print("  3. Risk Register        — Risk treatment tracking (Principle #8)")
    print("  4. KPI/KRI/KCI Metrics  — Situational awareness metrics (Principle #7)")


if __name__ == "__main__":
    main()
